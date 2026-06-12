#!/usr/bin/env python3
"""
Customer Ops Volume — source-of-truth refresh.

Pulls support volume across channels and writes three CSVs (the spreadsheet)
plus regenerates the Support Pulse dashboard, then commits + pushes.

Channels / metrics:
  Intercom  — success@ email, support@ email, chat  (+ avg first-response & resolution time)
  Linear    — CUS & REP issues created  (+ P0/P1 label counts)
  1 Metric  — % of CUS+REP issues created in period that were completed within 7 days
  Manual    — support_marco_escalations (prod/eng escalations; hand-entered, preserved)

Quo (phone) is intentionally out of scope for v1 — its API can't bulk-list calls.

Usage:
  python3 refresh_volume.py                 # incremental: refresh rolling 10d, redeploy
  python3 refresh_volume.py --backfill 90   # rebuild daily history from scratch (last 90d)
  python3 refresh_volume.py --no-deploy     # skip dashboard regen + git push

Writes (under support-pulse/):
  support_daily.csv  support_weekly.csv  support_monthly.csv
"""

import argparse, csv, datetime, json, os, re, subprocess, sys, time, urllib.request, urllib.error

REPO_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR  = os.path.join(REPO_DIR, "support-pulse")
GEN_SCRIPT = os.path.join(REPO_DIR, "generate_volume.py")
TODAY = datetime.date.today()

# ── env / auth (same pattern as refresh_data.py) ──────────────────────────────
def _load_env():
    env_path = os.path.join(REPO_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    os.environ.setdefault(k.strip(), v.strip())
_load_env()
INTERCOM_TOKEN = os.environ.get("INTERCOM_TOKEN", "")
LINEAR_TOKEN   = os.environ.get("LINEAR_TOKEN", "")

# ── source constants (verified 2026-06-09) ───────────────────────────────────
IC_SUCCESS_TAG = "Success@ Email"          # tag stamped on emails to success@
LINEAR_TEAM_KEYS = ["CUS", "REP"]          # Customer Success, Reporting
P0_LABEL, P1_LABEL = "P0", "P1"            # under "Customer Prioritization" group
SAME_WINDOW_DAYS = 7

# CSV column order. Leading "_"-cols are rollup helpers (kept in the file so
# weekly/monthly recompute exactly from daily without re-querying the APIs).
FETCH_COLS = [
    "ic_success_email", "ic_support_email", "ic_chat", "ic_total",
    "lin_cus_created", "lin_cus_p0", "lin_cus_p1",
    "lin_rep_created", "lin_rep_p0", "lin_rep_p1", "lin_total_created",
    "grand_total", "metric_created_completed_1wk_pct",
    "ic_avg_first_response_min", "ic_avg_resolution_hr",
]
HELPER_COLS = ["_ic_resp_sum_sec", "_ic_resp_n", "_ic_res_sum_sec", "_ic_res_n",
               "_metric_num", "_metric_den"]
MANUAL_COLS = ["support_marco_escalations"]   # never overwritten by the refresh
HEADER = (["period_label", "period_start", "period_end"]
          + FETCH_COLS[:4] + MANUAL_COLS + FETCH_COLS[4:] + HELPER_COLS)

# The single workbook is the human deliverable: clean columns only (no "_" helpers).
XLSX_PATH = os.path.join(OUT_DIR, "support_pulse.xlsx")
CLEAN_COLS = [c for c in HEADER if not c.startswith("_")]
PRETTY = {
    "period_label": "Period", "period_start": "Start", "period_end": "End",
    "ic_success_email": "IC success@", "ic_support_email": "IC support@",
    "ic_chat": "IC chat", "ic_total": "IC total",
    "support_marco_escalations": "Marco prod/eng escalations (manual)",
    "lin_cus_created": "CUS created", "lin_cus_p0": "CUS P0", "lin_cus_p1": "CUS P1",
    "lin_rep_created": "REP created", "lin_rep_p0": "REP P0", "lin_rep_p1": "REP P1",
    "lin_total_created": "Linear total", "grand_total": "Grand total",
    "metric_created_completed_1wk_pct": "1 Metric: created+completed ≤7d (%)",
    "ic_avg_first_response_min": "Avg 1st response (min)",
    "ic_avg_resolution_hr": "Avg resolution (hrs)",
}
SHEETS = [("Daily", "support_daily.csv"), ("Weekly", "support_weekly.csv"), ("Monthly", "support_monthly.csv")]

# Volume-by-type is stored long-format (one row per day×main×sub) so the dashboard
# can aggregate it across any selected date range and granularity.
CATEGORIES_PATH = os.path.join(OUT_DIR, "support_categories_daily.csv")
CATEGORIES_HEADER = ["date", "main", "sub", "count"]
USERTYPES_PATH = os.path.join(OUT_DIR, "support_usertypes_daily.csv")
USERTYPES_HEADER = ["date", "usertype", "count"]

# ── http helpers ──────────────────────────────────────────────────────────────
def http_post(url, headers, body, _retries=2):
    data = json.dumps(body).encode()
    req = urllib.request.Request(url, data=data, headers={**headers, "Content-Type": "application/json"})
    for attempt in range(_retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError:
            raise
        except (urllib.error.URLError, OSError):
            if attempt < _retries:
                time.sleep(2 ** attempt); continue
            raise

def _empty_day():
    return {c: 0 for c in FETCH_COLS + HELPER_COLS}

def strip_html(text):
    return re.sub(r"<[^>]+>", " ", text or "").strip()

# ── Volume-by-type taxonomy ───────────────────────────────────────────────────
# Each conversation is assigned to ONE (main, sub) by the FIRST matching rule, in
# the order below — so put specific/topical categories before channel/noise tags.
# Matching is case-insensitive substring against the conversation's Intercom tags.
# "Hzon"/"Hor" suffixes = Horizon; folded into the same buckets. Refine freely.
CATEGORY_RULES = [
    # main, sub, [keywords matched against tag text (lowercased)]
    ("Tickets", "Add tickets",        ["adding ticket", "add ticket"]),
    ("Tickets", "Edit tickets",       ["edit ticket"]),
    ("Tickets", "Delete tickets",     ["delete ticket"]),
    ("Tickets", "Find tickets",       ["find ticket"]),
    ("Tickets", "Move tickets",       ["move ticket"]),
    ("Tickets", "Download tickets",   ["download ticket"]),
    ("Tickets", "Digital tickets",    ["digital ticket"]),
    ("Tickets", "Other tickets",      ["ticket"]),
    ("Dispatch", "Dispatch help",     ["dispatch help"]),
    ("Dispatch", "Dispatch issues",   ["dispatch issue"]),
    ("Dispatch", "Dispatchers",       ["dispatcher"]),
    ("Dispatch", "Live map",          ["live map"]),
    ("Dispatch", "Dispatch (other)",  ["dispatch"]),
    ("Jobs", "Job cancellation",      ["job cancellation", "cancel"]),
    ("Jobs", "Job inquiries",         ["job inquir", "job inquiries"]),
    ("Jobs", "Job standby",           ["standby", "job-standby"]),
    ("Jobs", "Start job",             ["start job"]),
    ("Jobs", "Create order",          ["create order"]),
    ("Jobs", "Excess soil",           ["excess soil"]),
    ("Jobs", "Sites",                 ["sites"]),
    ("Billing & Rates", "Billing",    ["billing"]),
    ("Billing & Rates", "Invoicing",  ["invoic"]),
    ("Billing & Rates", "Fuel surcharge", ["fuel surcharge"]),
    ("Billing & Rates", "Rates",      ["rate"]),
    ("Billing & Rates", "Settlement/Payment", ["settlement", "payment"]),
    ("Integrations", "Apex",          ["apex"]),
    ("Integrations", "Axle",          ["axle"]),
    ("Integrations", "API",           ["api"]),
    ("Integrations", "Telematics",    ["telematics"]),
    ("Integrations", "Import/Export", ["import", "export"]),
    ("Integrations", "Integration (other)", ["integration"]),
    ("Reporting & Insights", "Reporting",   ["report", "gs report"]),
    ("Reporting & Insights", "Cycle times", ["cycle time"]),
    ("Reporting & Insights", "Insights",    ["insights"]),
    ("Reporting & Insights", "Data",        ["data"]),
    ("Mobile & App", "Mobile app",    ["mobile app"]),
    ("Mobile & App", "Driver app",    ["driver app"]),
    ("Mobile & App", "App storage",   ["app storage", "storage"]),
    ("Mobile & App", "Location/GPS",  ["location", "gps"]),
    ("Compliance", "Compliance",      ["compliance", "csa"]),
    ("Compliance", "Prevailing wage", ["prevailing wage"]),
    ("Compliance", "Hours",           ["checking hours", "clocking out"]),
    ("Account & Access", "Driver management", ["add driver", "move driver", "delete driver", "driver type", "driver day"]),
    ("Account & Access", "Vendor management", ["vendor", "add ven"]),
    ("Account & Access", "Login & access",    ["login", "get code", "access", "password", "sso"]),
    ("Account & Access", "Registration",      ["registration", "add client", "add customer", "add sub", "connected customer", "invite"]),
    ("Account & Access", "Equipment & trucks",["equipment", "truck"]),
    ("Account & Access", "Account support",   ["acc support", "account support"]),
    ("Feature Requests", "Feature request",   ["feature"]),
    ("Bugs", "Bug",                   ["bug"]),
    ("Other", "Looking for work",     ["looking for work"]),
    ("Other", "Lead",                 ["lead"]),
    ("Other", "General inquiries",    ["general inquir"]),
    ("Other", "Spam/Test",            ["spam", "test"]),
]
# Tags that are channel/routing/noise — not topical; ignored when categorizing.
CATEGORY_NOISE = {"email", "success@ email", "nar", "enterprise"}

def categorize(tags, text=""):
    """Return (main, sub) for one conversation from its tags + subject/body text.
    Tags are the strong signal; text reclaims conversations that were never tagged."""
    lowered = [t.lower() for t in tags if t and t.lower() not in CATEGORY_NOISE]
    body = (text or "").lower()
    for main, sub, kws in CATEGORY_RULES:
        for kw in kws:
            if any(kw in t for t in lowered) or kw in body:
                return main, sub
    return "Other", "Uncategorized"

# ── User-type (persona) taxonomy ──────────────────────────────────────────────
# "Which user types need the most support." Each conversation gets ONE primary
# persona: first from what the request is ABOUT (content keywords below), else
# from the requesting contact's assigned platform Roles, else "Unknown".
# Roles enum observed in Intercom contact "Roles" attribute.
USERTYPE_PRIORITY = ["driver", "dispatcher", "biller", "reporting", "foreman",
                     "manager", "companyAdmin", "itAdmin", "platformAdmin", "viewer"]
USERTYPE_LABEL = {
    "driver": "Driver", "dispatcher": "Dispatcher", "biller": "Biller",
    "reporting": "Reporting", "foreman": "Foreman", "manager": "Manager",
    "companyAdmin": "Company Admin", "itAdmin": "IT Admin",
    "platformAdmin": "Platform Admin", "viewer": "Viewer", "Unknown": "Unknown",
}
USERTYPE_KW = {
    "driver":      ["driver app", "my truck", "check in", "check-in", "clock in", "clock out",
                    "clocking", "geofence", "log in to the app", "can't log into the app", "haul"],
    "dispatcher":  ["dispatch", "assign truck", "assign driver", "schedule", "send drivers",
                    "job board", "move ticket", "find ticket", "live map"],
    "biller":      ["invoice", "invoicing", "billing", "settlement", "payment", "fuel surcharge", "rate"],
    "reporting":   ["report", "dashboard", "cycle time", "insights", "export data", "prevailing wage"],
    "foreman":     ["foreman", "field", "on site", "jobsite", "job site"],
    "companyAdmin":["add driver", "add vendor", "add user", "add equipment", "add truck",
                    "register", "registration", "permission", "set up account", "onboard", "invite user"],
}

def normalize_roles(raw):
    """Parse the contact 'Roles' attribute value into known role tokens."""
    if not raw:
        return []
    toks = [t.strip() for t in str(raw).replace(";", ",").split(",")]
    return [t for t in toks if t in USERTYPE_PRIORITY]

def user_type(text, roles):
    """One primary persona for a conversation: content first, then assigned role, else Unknown."""
    body = (text or "").lower()
    hits = [r for r in USERTYPE_PRIORITY if r in USERTYPE_KW
            and any(kw in body for kw in USERTYPE_KW[r])]
    if hits:
        return USERTYPE_LABEL[min(hits, key=USERTYPE_PRIORITY.index)]
    roles = [r for r in roles if r in USERTYPE_PRIORITY]
    if roles:
        return USERTYPE_LABEL[min(roles, key=USERTYPE_PRIORITY.index)]
    return "Unknown"

# Contact→roles resolver with a persistent cache (kept OUT of the public repo).
CONTACT_CACHE_PATH = os.path.join(REPO_DIR, ".contact_roles_cache.json")

def load_contact_cache():
    if os.path.exists(CONTACT_CACHE_PATH):
        try:
            with open(CONTACT_CACHE_PATH) as f:
                return json.load(f)
        except (ValueError, OSError):
            return {}
    return {}

def save_contact_cache(cache):
    try:
        with open(CONTACT_CACHE_PATH, "w") as f:
            json.dump(cache, f)
    except OSError:
        pass

def resolve_roles(contact_id, cache, headers):
    """Return the contact's role tokens, using/refreshing the cache. [] on miss/error."""
    if not contact_id:
        return []
    if contact_id in cache:
        return cache[contact_id]
    try:
        req = urllib.request.Request(f"https://api.intercom.io/contacts/{contact_id}", headers=headers)
        with urllib.request.urlopen(req, timeout=30) as r:
            ca = (json.loads(r.read()).get("custom_attributes") or {})
        roles = normalize_roles(ca.get("Roles") or ca.get("Role") or "")
    except (urllib.error.URLError, OSError, ValueError):
        roles = []
    cache[contact_id] = roles
    return roles

# ── Intercom ──────────────────────────────────────────────────────────────────
def fetch_intercom(days, by_day, cat_by_day=None, ut_by_day=None, contact_cache=None):
    """Page all conversations created in [now-days, now]; bucket counts + timing per day.
    Also buckets volume-by-type (cat_by_day) and volume-by-user-type (ut_by_day)."""
    since = int((datetime.datetime.now() - datetime.timedelta(days=days)).timestamp())
    headers = {"Authorization": f"Bearer {INTERCOM_TOKEN}", "Accept": "application/json",
               "Intercom-Version": "2.11"}
    if contact_cache is None:
        contact_cache = {}
    starting_after, n = None, 0
    while True:
        body = {"query": {"operator": "AND", "value": [
                    {"field": "created_at", "operator": ">", "value": since}]},
                "pagination": {"per_page": 150}}
        if starting_after:
            body["pagination"]["starting_after"] = starting_after
        data = http_post("https://api.intercom.io/conversations/search", headers, body)
        convs = data.get("conversations", [])
        if not convs:
            break
        for c in convs:
            n += 1
            d = datetime.datetime.fromtimestamp(c.get("created_at", 0)).date().isoformat()
            row = by_day.setdefault(d, _empty_day())
            src = c.get("source", {})
            stype = src.get("type")
            tags = [t.get("name", "") for t in c.get("tags", {}).get("tags", [])]
            if stype == "conversation":
                row["ic_chat"] += 1
            elif stype == "email":
                if IC_SUCCESS_TAG in tags:
                    row["ic_success_email"] += 1
                else:
                    row["ic_support_email"] += 1
            else:
                continue  # ignore admin_initiated / other source types for channel volume
            text = (src.get("subject") or "") + " " + strip_html(src.get("body") or "")
            # volume-by-type: assign this conversation to one (main, sub)
            if cat_by_day is not None:
                main, sub = categorize(tags, text)
                cat_by_day.setdefault(d, {}).setdefault(main, {}).setdefault(sub, 0)
                cat_by_day[d][main][sub] += 1
            # volume-by-user-type: one primary persona (content, then contact role)
            if ut_by_day is not None:
                cons = c.get("contacts", {}).get("contacts", [])
                roles = resolve_roles(cons[0]["id"], contact_cache, headers) if cons else []
                ut = user_type(text, roles)
                ut_by_day.setdefault(d, {}).setdefault(ut, 0)
                ut_by_day[d][ut] += 1
            # timing (only present once a human replied / closed)
            st = c.get("statistics") or {}
            ttr = st.get("time_to_admin_reply")
            if isinstance(ttr, (int, float)) and ttr >= 0:
                row["_ic_resp_sum_sec"] += ttr; row["_ic_resp_n"] += 1
            ttc = st.get("time_to_last_close")
            if isinstance(ttc, (int, float)) and ttc >= 0:
                row["_ic_res_sum_sec"] += ttc; row["_ic_res_n"] += 1
        pages = data.get("pages", {})
        nxt = pages.get("next")
        starting_after = nxt.get("starting_after") if isinstance(nxt, dict) else None
        if not starting_after:
            break
        time.sleep(0.12)
    print(f"  Intercom: scanned {n} conversations over {days}d")

# ── Linear ──────────────────────────────────────────────────────────────────
LINEAR_GQL = """
query Created($after: String, $since: DateTimeOrDuration!) {
  issues(
    filter: { team: { key: { in: ["CUS","REP"] } }, createdAt: { gte: $since } }
    first: 100
    after: $after
    orderBy: createdAt
  ) {
    pageInfo { hasNextPage endCursor }
    nodes {
      identifier createdAt completedAt
      team { key }
      labels { nodes { name } }
    }
  }
}
"""

def fetch_linear(days, by_day):
    since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    headers = {"Authorization": LINEAR_TOKEN, "Content-Type": "application/json"}
    cursor, n = None, 0
    while True:
        variables = {"since": since}
        if cursor:
            variables["after"] = cursor
        data = http_post("https://api.linear.app/graphql", headers,
                         {"query": LINEAR_GQL, "variables": variables})
        if "errors" in data:
            print("  Linear GQL error:", json.dumps(data["errors"])[:300]); break
        page = data.get("data", {}).get("issues", {})
        for node in page.get("nodes", []):
            n += 1
            created = node.get("createdAt", "")[:10]
            if not created:
                continue
            row = by_day.setdefault(created, _empty_day())
            key = (node.get("team") or {}).get("key", "")
            labels = [l.get("name", "") for l in (node.get("labels") or {}).get("nodes", [])]
            pre = "lin_cus" if key == "CUS" else "lin_rep" if key == "REP" else None
            if not pre:
                continue
            row[f"{pre}_created"] += 1
            row["lin_total_created"] += 1
            if P0_LABEL in labels: row[f"{pre}_p0"] += 1
            if P1_LABEL in labels: row[f"{pre}_p1"] += 1
            # 1-metric: created & completed within SAME_WINDOW_DAYS
            row["_metric_den"] += 1
            comp = node.get("completedAt")
            if comp:
                try:
                    c_dt = datetime.datetime.fromisoformat(node["createdAt"].replace("Z", "+00:00"))
                    d_dt = datetime.datetime.fromisoformat(comp.replace("Z", "+00:00"))
                    if (d_dt - c_dt).total_seconds() <= SAME_WINDOW_DAYS * 86400:
                        row["_metric_num"] += 1
                except ValueError:
                    pass
        pi = page.get("pageInfo", {})
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor"); time.sleep(0.15)
    print(f"  Linear: scanned {n} CUS/REP issues created over {days}d")

# ── derived columns ───────────────────────────────────────────────────────────
def finalize_day(row):
    row["ic_total"] = row["ic_success_email"] + row["ic_support_email"] + row["ic_chat"]
    row["grand_total"] = row["ic_total"] + row["lin_total_created"]
    row["ic_avg_first_response_min"] = round(row["_ic_resp_sum_sec"] / row["_ic_resp_n"] / 60, 1) if row["_ic_resp_n"] else ""
    row["ic_avg_resolution_hr"]      = round(row["_ic_res_sum_sec"] / row["_ic_res_n"] / 3600, 1) if row["_ic_res_n"] else ""
    row["metric_created_completed_1wk_pct"] = round(100 * row["_metric_num"] / row["_metric_den"], 1) if row["_metric_den"] else ""

# ── CSV read / write (preserves manual columns) ───────────────────────────────
def read_csv(path):
    if not os.path.exists(path):
        return {}
    out = {}
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            out[r["period_label"]] = r
    return out

def write_csv(path, rows):
    rows = sorted(rows, key=lambda r: r["period_start"])
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADER)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in HEADER})

def read_xlsx_manual():
    """Manual columns are edited in the workbook — read them back so edits persist.
    Returns {sheet_name: {period_label: {manual_col: value}}}. Empty if no workbook yet."""
    out = {name: {} for name, _ in SHEETS}
    if not os.path.exists(XLSX_PATH):
        return out
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    for name, _ in SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        try:
            hdr = list(next(rows))
        except StopIteration:
            continue
        idx = {h: i for i, h in enumerate(hdr)}
        if "Period" not in idx:
            continue
        for r in rows:
            label = r[idx["Period"]]
            if label is None:
                continue
            rec = {}
            for mc in MANUAL_COLS:
                pi = idx.get(PRETTY[mc])
                if pi is not None and r[pi] not in (None, ""):
                    rec[mc] = r[pi]
            if rec:
                out[name][str(label)] = rec
    wb.close()
    return out

def write_xlsx():
    """Write the single deliverable workbook with Daily / Weekly / Monthly tabs (clean columns)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="132732")
    head_font = Font(bold=True, color="FFE500")
    for name, csv_name in SHEETS:
        ws = wb.create_sheet(name)
        ws.append([PRETTY[c] for c in CLEAN_COLS])
        for cell in ws[1]:
            cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in read_csv_list(os.path.join(OUT_DIR, csv_name)):
            ws.append([_xnum(r.get(c, "")) for c in CLEAN_COLS])
        ws.freeze_panes = "B2"
        for i, c in enumerate(CLEAN_COLS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                12 if c.startswith(("ic_", "lin_", "grand", "metric", "period_s", "period_e")) else \
                (22 if c in ("support_marco_escalations", "metric_created_completed_1wk_pct") else 14)
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(CLEAN_COLS))}1"
    wb.save(XLSX_PATH)
    print(f"  → {XLSX_PATH} ({len(SHEETS)} tabs)")

def _xnum(v):
    if v in (None, ""):
        return None
    try:
        f = float(v); return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return v

def read_csv_list(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))

def read_categories():
    """Load the long-format categories CSV into {date: {main: {sub: count}}}."""
    out = {}
    if not os.path.exists(CATEGORIES_PATH):
        return out
    with open(CATEGORIES_PATH, newline="") as f:
        for r in csv.DictReader(f):
            try:
                cnt = int(float(r["count"]))
            except (ValueError, KeyError):
                continue
            out.setdefault(r["date"], {}).setdefault(r["main"], {})[r["sub"]] = cnt
    return out

def write_categories(cat_by_day):
    rows = []
    for d in sorted(cat_by_day):
        for main in sorted(cat_by_day[d]):
            for sub, cnt in sorted(cat_by_day[d][main].items()):
                rows.append({"date": d, "main": main, "sub": sub, "count": cnt})
    with open(CATEGORIES_PATH, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CATEGORIES_HEADER)
        w.writeheader(); w.writerows(rows)
    return len(rows)

def read_usertypes():
    out = {}
    if not os.path.exists(USERTYPES_PATH):
        return out
    with open(USERTYPES_PATH, newline="") as f:
        for r in csv.DictReader(f):
            try:
                out.setdefault(r["date"], {})[r["usertype"]] = int(float(r["count"]))
            except (ValueError, KeyError):
                continue
    return out

def write_usertypes(ut_by_day):
    rows = []
    for d in sorted(ut_by_day):
        for ut, cnt in sorted(ut_by_day[d].items(), key=lambda x: -x[1]):
            rows.append({"date": d, "usertype": ut, "count": cnt})
    with open(USERTYPES_PATH, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=USERTYPES_HEADER)
        w.writeheader(); w.writerows(rows)
    return len(rows)

def build_daily_rows(by_day, existing):
    rows = []
    for d in sorted(by_day):
        finalize_day(by_day[d])
        rec = {"period_label": d, "period_start": d, "period_end": d}
        rec.update({k: by_day[d][k] for k in FETCH_COLS + HELPER_COLS})
        # preserve hand-entered manual columns
        for mc in MANUAL_COLS:
            rec[mc] = (existing.get(d) or {}).get(mc, "")
        rows.append(rec)
    return rows

# ── weekly / monthly rollups (recomputed from daily; exact via helper sums) ───
def _agg(rows, key_fn, label_fn, span_fn, existing_manual):
    buckets = {}
    for r in rows:
        d = datetime.date.fromisoformat(r["period_start"])
        k = key_fn(d)
        b = buckets.setdefault(k, _empty_day())
        for c in ["ic_success_email", "ic_support_email", "ic_chat",
                  "lin_cus_created", "lin_cus_p0", "lin_cus_p1",
                  "lin_rep_created", "lin_rep_p0", "lin_rep_p1", "lin_total_created"] + HELPER_COLS:
            b[c] += int(float(r.get(c) or 0))
    out = []
    for k, b in buckets.items():
        finalize_day(b)
        start, end = span_fn(k)
        label = label_fn(k)
        rec = {"period_label": label, "period_start": start.isoformat(), "period_end": end.isoformat()}
        rec.update({c: b[c] for c in FETCH_COLS + HELPER_COLS})
        for mc in MANUAL_COLS:
            rec[mc] = (existing_manual.get(label) or {}).get(mc, "")
        out.append(rec)
    return out

def iso_week_key(d):  # (iso_year, iso_week)
    iy, iw, _ = d.isocalendar(); return (iy, iw)
def iso_week_span(k):
    iy, iw = k
    start = datetime.date.fromisocalendar(iy, iw, 1)
    return start, start + datetime.timedelta(days=6)
def iso_week_label(k):
    iy, iw = k; return f"{iy}-W{iw:02d}"

def month_key(d):   return (d.year, d.month)
def month_span(k):
    y, m = k
    start = datetime.date(y, m, 1)
    end = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1))
    return start, end
def month_label(k): return f"{k[0]}-{k[1]:02d}"

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--backfill", type=int, metavar="N",
                    help="rebuild daily history from scratch for the last N days")
    ap.add_argument("--deploy", action="store_true", default=True)
    ap.add_argument("--no-deploy", dest="deploy", action="store_false")
    args = ap.parse_args()

    os.makedirs(OUT_DIR, exist_ok=True)
    daily_path = os.path.join(OUT_DIR, "support_daily.csv")
    existing = read_csv(daily_path)

    # Manual columns are edited in the workbook — read them back (xlsx wins over CSV).
    xman = read_xlsx_manual()
    def manual_for(sheet, csv_rows):
        out = {}
        for label, row in csv_rows.items():
            rec = {mc: row.get(mc, "") for mc in MANUAL_COLS if row.get(mc, "") not in (None, "")}
            if rec:
                out[label] = rec
        out.update(xman.get(sheet, {}))
        return out
    man_daily = manual_for("Daily", existing)

    window = args.backfill if args.backfill else 10   # incremental refreshes a rolling 10d
    print(f"Fetching {'backfill ' if args.backfill else 'incremental '}{window}d…")
    by_day, cat_by_day, ut_by_day = {}, {}, {}
    contact_cache = load_contact_cache()
    fetch_intercom(window, by_day, cat_by_day, ut_by_day, contact_cache)
    save_contact_cache(contact_cache)
    fetch_linear(window, by_day)
    fresh = build_daily_rows(by_day, man_daily)

    if args.backfill:
        merged = {r["period_label"]: r for r in fresh}              # full rebuild of window
        for lbl, r in existing.items():                            # keep older days outside window
            if lbl not in merged and lbl < min(by_day):
                merged[lbl] = r
    else:
        merged = dict(existing)                                     # upsert the rolling window
        merged.update({r["period_label"]: r for r in fresh})
    daily_rows = list(merged.values())
    write_csv(daily_path, daily_rows)
    print(f"  → {daily_path} ({len(daily_rows)} days)")

    # rollups (preserve manual values — workbook edits win, CSV as fallback)
    man_weekly  = manual_for("Weekly",  read_csv(os.path.join(OUT_DIR, "support_weekly.csv")))
    man_monthly = manual_for("Monthly", read_csv(os.path.join(OUT_DIR, "support_monthly.csv")))
    weekly  = _agg(daily_rows, iso_week_key, iso_week_label, iso_week_span, man_weekly)
    monthly = _agg(daily_rows, month_key, month_label, month_span, man_monthly)
    write_csv(os.path.join(OUT_DIR, "support_weekly.csv"), weekly)
    write_csv(os.path.join(OUT_DIR, "support_monthly.csv"), monthly)
    print(f"  → support_weekly.csv ({len(weekly)} weeks), support_monthly.csv ({len(monthly)} months)")

    # volume-by-type: upsert the same window into the long-format categories file
    cat_existing = read_categories()
    if args.backfill:
        cutoff = min(cat_by_day) if cat_by_day else None
        merged_cat = dict(cat_by_day)
        for d, v in cat_existing.items():
            if cutoff and d < cutoff and d not in merged_cat:
                merged_cat[d] = v
    else:
        merged_cat = dict(cat_existing)
        merged_cat.update(cat_by_day)   # refreshed days overwrite
    ncat = write_categories(merged_cat)
    print(f"  → support_categories_daily.csv ({ncat} rows, {len(merged_cat)} days)")

    # volume-by-user-type: upsert the same window into the long-format usertypes file
    ut_existing = read_usertypes()
    if args.backfill:
        cutoff = min(ut_by_day) if ut_by_day else None
        merged_ut = dict(ut_by_day)
        for d, v in ut_existing.items():
            if cutoff and d < cutoff and d not in merged_ut:
                merged_ut[d] = v
    else:
        merged_ut = dict(ut_existing)
        merged_ut.update(ut_by_day)
    nut = write_usertypes(merged_ut)
    print(f"  → support_usertypes_daily.csv ({nut} rows, {len(merged_ut)} days)")

    write_xlsx()   # single deliverable workbook (3 tabs)

    if args.deploy:
        deploy()

def deploy():
    print("Regenerating dashboard…")
    if os.path.exists(GEN_SCRIPT):
        r = subprocess.run([sys.executable, GEN_SCRIPT], capture_output=True, text=True)
        if r.returncode != 0:
            print("  dashboard ERROR:", r.stderr[-500:]); sys.exit(1)
        print(" ", (r.stdout or "").strip())
    print("Committing + pushing…")
    subprocess.run(["git", "-C", REPO_DIR, "add", "support-pulse", "index.html"], check=True)
    diff = subprocess.run(["git", "-C", REPO_DIR, "diff", "--staged", "--quiet"])
    if diff.returncode == 0:
        print("  no changes to push."); return
    subprocess.run(["git", "-C", REPO_DIR, "commit", "-m", f"Support Pulse refresh {TODAY.isoformat()}"], check=True)
    subprocess.run(["git", "-C", REPO_DIR, "push"], check=True)
    print("  pushed — GitHub Pages updates in ~30s")

if __name__ == "__main__":
    main()
